"""
Created on Friday Aug 16 2019

@author: carlos.delaguardia
"""
from  openpyxl import load_workbook
from math import floor

FILE_PATH = 'Recursos.xlsx'
HOJA= 'Hoja1'
ALM_CT=10
libro=load_workbook(FILE_PATH, read_only=False)
hoja=libro[HOJA]

E009= hoja['A2'].value

NecesidadReal_Total=0
Repartido1=0
Repartido2=0
Queda1=0
Queda2=0
NEcesidad_Total=0
Ceros=0
Maximo=0

j=2
for k in range (ALM_CT):
    NEcesidad_Total=NEcesidad_Total+(hoja.cell(row=j, column=5).value)
    j=j+1
